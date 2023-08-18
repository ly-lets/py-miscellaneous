import os
import sys
from docx import Document
from openpyxl import load_workbook
import PyPDF2
import win32com.client
import csv
import time

failedOperations = []
output_level = 1

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def open_word_in_msword(file_path, msword):
    try:
        curlevel = output_level + 2
        # word.visible = False
        wb = msword.Documents.Open(file_path)
        doc = msword.ActiveDocument
        # Get word count
        word_count = doc.ComputeStatistics(win32com.client.constants.wdStatisticWords)
        # print(doc.Range().Text)
        print( f'{bcolors.ENDC}{file_path}  {bcolors.HEADER} {word_count}\n')
        # print(len(doc.Range().Text))
        return word_count, os.path.basename(file_path)
    except Exception as e:
        failedOperations.append([os.path.basename(file_path),e.__class__.__name__])
        print( f"{bcolors.FAIL} Warning:expcetion occurred : {e.__class__.__name__}")
    finally:
        doc.Close(False)

def count_words_in_docx(file_path):
    document = Document(file_path)
    word_count = 0
    for paragraph in document.paragraphs:
        word_count += len(paragraph.text.split())
    return word_count, os.path.basename(file_path)


def count_words_in_xlsx(file_path):
    workbook = load_workbook(file_path)
    word_count = 0
    for sheet in workbook.sheetnames:
        for row in workbook[sheet].iter_rows(values_only=True):
            for cell_value in row:
                if cell_value is not None:
                    word_count += len(str(cell_value).split())
    print(f'{bcolors.ENDC}{file_path} {bcolors.HEADER} {word_count}\n')
    return word_count, os.path.basename(file_path)


def count_words_in_pdf(file_path):
    with open(file_path, 'rb') as file:
        pdf = PyPDF2.PdfReader(file)
        word_count = 0
        for page_num in range(len(pdf.pages)):
            page = pdf.pages[page_num]
            word_count += len(page.extract_text().split())
    return word_count, os.path.basename(file_path)


def calculate_total_word_count(directory):
    curlevel = output_level + 1
    print( f"{bcolors.HEADER} Preparing MS word application")
    word =  win32com.client.gencache.EnsureDispatch("Word.Application")

    word_count = 0
    counted_files = []
    print( f'Working at {bcolors.OKGREEN} ---> {directory}')
    for root, dirs, files in os.walk(directory):
        for file_name in files:
            file_path = os.path.join(root, file_name)
            if file_name.startswith('~$'): continue
            if file_name.endswith('.doc') or file_name.endswith('.docx'):
                count, file_name = open_word_in_msword(file_path, word)
                word_count += count
                counted_files.append((file_name, count))
            # elif file_name.endswith('.docx'):
            #     count, file_name = count_words_in_docx(file_path)
            #     word_count += count
            #     counted_files.append((file_name, count))
            elif file_name.endswith('.xlsx'):
                count, file_name = count_words_in_xlsx(file_path)
                word_count += count
                counted_files.append((file_name, count))
            # elif file_name.endswith('.pdf'):
            #     count, file_name = count_words_in_pdf(file_path)
            #     word_count += count
            #     counted_files.append((file_name, count))
    counted_files_with_total = [
        (f"{file_name}", count) for file_name, count in counted_files]
    
    word.Quit()

    return word_count, counted_files_with_total


def main():
    curlevel = output_level
    target_directory = ''
    if (len(sys.argv) < 2):
        target_directory = input("Enter the full path to the target directory: ")
    else:
        target_directory = sys.argv[1]
    # target_directory = re.escape("E:\workspace\python\test")

    total_word_count, counted_files = calculate_total_word_count(
        target_directory)
    
    output_file = "counted_files.txt"
    with open(output_file, "w", encoding="utf-8") as file:
        for file_name, count in counted_files:
            file.write(f"{file_name} , {count}\n")

    print( f"{bcolors.WARNING}Total Word Count: {bcolors.OKGREEN} {total_word_count}" )
    print( f"{bcolors.WARNING}Counted files saved to{bcolors.OKGREEN} {output_file}")
    print(f"{bcolors.ENDC}")

if __name__ == "__main__":
    main()
